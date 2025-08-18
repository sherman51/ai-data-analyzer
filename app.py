import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from itertools import cycle
from datetime import datetime
from zoneinfo import ZoneInfo

# ------------------------ UI CONFIGURATION ------------------------
st.set_page_config(page_title="Master Pick Ticket Generator", layout="wide")
st.title("📦 Master Pick Ticket Generator – ")

# ------------------------ FILE UPLOADS ------------------------
st.sidebar.header("📂 Upload Input Files")
picking_pool_file = st.sidebar.file_uploader("Upload Picking Pool Excel file", type=["xlsx"])
sku_master_file = st.sidebar.file_uploader("Upload SKU Master Excel file", type=["xlsx"])

# ------------------------ HELPER FUNCTIONS ------------------------

def map_location_to_zone(location: str) -> str:
    """
    Map warehouse locations into Zones based on the first 2-digit section after 'A-'.
    Example:
      A-01 to A-05 -> Zone 1
      A-06 to A-10 -> Zone 2
      A-11 to A-15 -> Zone 3
      A-16 to A-20 -> Zone 4
    """
    if not isinstance(location, str):
        return "Unknown"

    location = location.strip().upper()

    # Extract the first number after 'A-'
    match = re.match(r"A-(\d{2})", location)
    if match:
        num = int(match.group(1))
        if 1 <= num <= 5:
            return "Zone 1"
        elif 6 <= num <= 10:
            return "Zone 2"
        elif 11 <= num <= 15:
            return "Zone 3"
        elif 16 <= num <= 20:
            return "Zone 4"
        else:
            return "Other A Zone"

    if location.startswith("SOFT-"):
        return "Soft"

    return "Unknown"


def calculate_carton_info(row):
    pq = row.get('PickingQty', 0) or 0
    qpc = row.get('Qty per Carton', 0) or 0
    iv = row.get('Item Vol', 0) or 0
    qpco = row.get('Qty Commercial Box', 0) or 0

    if pq == 0 or qpc == 0 or iv == 0:
        return pd.Series({'CartonCount': None, 'CartonDescription': 'Invalid'})

    cartons = int(pq // qpc)
    loose = pq % qpc

    carton_sizes = [
        (1200, "1XS"),
        (6000, "1S"),
        (12000, "1Rectangle"),
        (48000, "1L"),
        (float('inf'), "Too Big")
    ]

    if loose > 0:
        looseVol = loose / qpco * iv
        looseBox = next(name for max_vol, name in carton_sizes if looseVol <= max_vol)
        desc = f"{cartons} Commercial Carton + {looseBox}" if cartons > 0 else looseBox
        totalC = cartons + 1
    else:
        desc = f"{cartons} Commercial Carton"
        totalC = cartons

    return pd.Series({'CartonCount': totalC, 'CartonDescription': desc})


def classify_gi(volume):
    if pd.isna(volume):
        return 'Unknown'
    if volume < 35000:
        return 'Bin'
    elif volume < 248500:
        return 'Layer'
    else:
        return 'Pick by Orders'


def load_data(picking_pool_file, sku_master_file):
    picking_pool = pd.read_excel(picking_pool_file)
    sku_master = pd.read_excel(sku_master_file)

    picking_pool['DeliveryDate'] = pd.to_datetime(picking_pool['DeliveryDate'], errors='coerce')
    picking_pool = picking_pool[picking_pool['DeliveryDate'].notna()]
    picking_pool['Delivery Date'] = picking_pool['DeliveryDate'].dt.strftime("%Y-%m-%d")

    return picking_pool, sku_master


def filter_picking_pool(df):
    # Normalize LocationType column
    df['LocationType'] = df['LocationType'].astype(str).str.strip().str.lower()

    # --- Identify IssueNos that have any 'storage' line ---
    storage_gis = df[df['LocationType'] == 'storage']['IssueNo'].unique()

    # --- Remove entire GI if it contains storage lines ---
    df = df[~df['IssueNo'].isin(storage_gis)]

    # Add ZoneMapped column
    df['ZoneMapped'] = df['Location'].apply(map_location_to_zone)

    return df


def apply_delivery_date_filter(df, delivery_range):
    if isinstance(delivery_range, tuple) and len(delivery_range) == 2:
        start, end = pd.to_datetime(delivery_range[0]), pd.to_datetime(delivery_range[1])
        return df[(df['DeliveryDate'] >= start) & (df['DeliveryDate'] <= end)]
    return df


def merge_and_clean(df, sku_master):
    merged_check = df.merge(sku_master, how='left', left_on='SKU', right_on='SKU Code')
    missing_issues = merged_check[
        merged_check['Qty Commercial Box'].isna() |
        merged_check['Qty per Carton'].isna() |
        merged_check['Item Vol'].isna()
    ]['IssueNo'].unique()
    df = df[~df['IssueNo'].isin(missing_issues)]

    df = df.merge(sku_master, how='left', left_on='SKU', right_on='SKU Code')
    df['PickingQty'] = df['PickingQty'].fillna(0)
    df['Item Vol'] = df['Item Vol'].fillna(0)
    df['Qty Commercial Box'] = df['Qty Commercial Box'].replace(0, 1).fillna(1)
    df['Qty per Carton'] = df['Qty per Carton'].replace(0, 1).fillna(1)
    df['Total Item Vol'] = (df['PickingQty'] / df['Qty Commercial Box']) * df['Item Vol']
    return df


def add_line_count(df):
    line_counts = df.groupby('IssueNo').size()
    df = df.copy()
    df['Line Count'] = df['IssueNo'].map(line_counts)
    return df


def classify_and_assign(df):
    df = df.merge(df.groupby('IssueNo')['Total Item Vol'].sum().rename('Total GI Vol'), on='IssueNo')
    df = add_line_count(df)
    df['GI Class'] = df['Total GI Vol'].apply(classify_gi)
    df = df[df['Total GI Vol'] <= 248500]

    bin_gi_issues = df[df['GI Class'] == 'Bin']['IssueNo'].unique()
    layer_gi_issues = df[df['GI Class'] == 'Layer']['IssueNo'].unique()

    bin_no_mapping = {issue: f"Bin{str(i+1).zfill(3)}" for i, issue in enumerate(sorted(bin_gi_issues))}
    layer_no_mapping = {issue: f"Layer{str(i+1).zfill(3)}" for i, issue in enumerate(sorted(layer_gi_issues))}

    df['BinNo'] = df['IssueNo'].map(bin_no_mapping)
    df['LayerNo'] = df['IssueNo'].map(layer_no_mapping)
    df['Type'] = df.apply(lambda row: row['BinNo'] if row['GI Class'] == 'Bin' else row['LayerNo'], axis=1)

    return df


def assign_job_numbers_with_scenarios(df):
    """
    Assign Job No with updated logic:
    - All lines of the same IssueNo (GI) always go into the same Job No
    - Single-line GIs grouped by ZoneMapped + DeliveryDate, max 4 GIs per job
    - Multi-line GIs use scenarios (2 bins + 1 layer, etc.)
    """
    df = df.copy()

    # Work only on unique IssueNos
    gi_info = df[['IssueNo', 'ShipToName', 'Line Count', 'GI Class', 'DeliveryDate', 'ZoneMapped']].drop_duplicates()

    job_no_counter = 1
    gi_info['Job No'] = None

    # --- SINGLE-LINE GIs: Group by ZoneMapped + DeliveryDate ---
    single_line = gi_info[gi_info['Line Count'] == 1].copy()
    for (zone, delivery_date), group in single_line.groupby(['ZoneMapped', 'DeliveryDate']):
        issues = group['IssueNo'].tolist()
        chunks = [issues[i:i+4] for i in range(0, len(issues), 4)]  # max 4 per job
        for chunk in chunks:
            job_no_str = f"Job{str(job_no_counter).zfill(3)}"
            gi_info.loc[gi_info['IssueNo'].isin(chunk), 'Job No'] = job_no_str
            job_no_counter += 1

    # --- MULTI-LINE GIs ---
    multi_line = gi_info[gi_info['Line Count'] > 1].copy()
    for delivery_date, group in multi_line.groupby('DeliveryDate'):
        bins = group[group['GI Class'] == 'Bin']['IssueNo'].tolist()
        layers = group[group['GI Class'] == 'Layer']['IssueNo'].tolist()

        # Scenario 1: 2 Bins + 1 Layer
        while len(bins) >= 2 and len(layers) >= 1:
            job_issues = bins[:2] + [layers[0]]
            job_no_str = f"Job{str(job_no_counter).zfill(3)}"
            gi_info.loc[gi_info['IssueNo'].isin(job_issues), 'Job No'] = job_no_str
            job_no_counter += 1
            bins = bins[2:]
            layers = layers[1:]

        # Scenario 2: Only Bins (3–4 per job)
        while len(bins) >= 3:
            group_size = 4 if len(bins) >= 4 else 3
            job_issues = bins[:group_size]
            job_no_str = f"Job{str(job_no_counter).zfill(3)}"
            gi_info.loc[gi_info['IssueNo'].isin(job_issues), 'Job No'] = job_no_str
            job_no_counter += 1
            bins = bins[group_size:]

        # Scenario 3: Only Layers (2–3 per job)
        while len(layers) >= 2:
            group_size = 3 if len(layers) >= 3 else 2
            job_issues = layers[:group_size]
            job_no_str = f"Job{str(job_no_counter).zfill(3)}"
            gi_info.loc[gi_info['IssueNo'].isin(job_issues), 'Job No'] = job_no_str
            job_no_counter += 1
            layers = layers[group_size:]

        # Remaining singles
        for issue in bins + layers:
            job_no_str = f"Job{str(job_no_counter).zfill(3)}"
            gi_info.loc[gi_info['IssueNo'] == issue, 'Job No'] = job_no_str
            job_no_counter += 1

    # --- Merge Job No back to full df ---
    df = df.merge(gi_info[['IssueNo', 'Job No']], on='IssueNo', how='left')

    # Sanity check: ensure no GI split into multiple jobs
    check = df.groupby('IssueNo')['Job No'].nunique()
    bad_gis = check[check > 1]
    if not bad_gis.empty:
        st.error(f"❌ These IssueNos were split across multiple jobs: {', '.join(map(str, bad_gis.index))}")

    return df


def finalize_output(df, gi_type):
    if gi_type == "Single-line":
        df = df[df['Line Count'] == 1]
    elif gi_type == "Multi-line":
        df = df[df['Line Count'] > 1]

    df[['CartonCount', 'CartonDescription']] = df.apply(calculate_carton_info, axis=1)
    df['Batch No'] = df.get('StorageLocation')
    df['Commercial Box Count'] = df['PickingQty'] / df['Qty Commercial Box']

    return df[[
        'IssueNo', 'SKU', 'Location_x', 'SKUDescription', 'Batch No', 'PickingQty',
        'Commercial Box Count', 'Delivery Date', 'ShipToName',
        'Type', 'Job No', 'CartonDescription', 'ZoneMapped'
    ]].drop_duplicates()


def export_to_excel(output_df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False, sheet_name='Master Pick Ticket')
        worksheet = writer.sheets['Master Pick Ticket']

        for col_idx, column_cells in enumerate(worksheet.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        sku_batch_group = output_df.groupby('SKU')['Batch No'].nunique()
        skus_with_diff_batch = sku_batch_group[sku_batch_group > 1].index

        color_cycle = cycle([
            PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid"),
            PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid"),
            PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"),
            PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid"),
            PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),
        ])

        sku_color_map = {sku: next(color_cycle) for sku in skus_with_diff_batch}

        for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=worksheet.max_column):
            sku_cell = row[1]
            if sku_cell.value in sku_color_map:
                for cell in row:
                    cell.fill = sku_color_map[sku_cell.value]

        output.seek(0)

    return output

# ------------------------ MAIN LOGIC ------------------------

def main():
    try:
        picking_pool, sku_master = load_data(picking_pool_file, sku_master_file)

        picking_pool = filter_picking_pool(picking_pool)

        gi_type = st.sidebar.radio("Filter by GI Type", ("All", "Single-line", "Multi-line"))

        min_date, max_date = picking_pool['DeliveryDate'].min(), picking_pool['DeliveryDate'].max()
        delivery_range = st.sidebar.date_input("🗕️ Filter by Delivery Date", (min_date, max_date), min_value=min_date, max_value=max_date)

        picking_pool = apply_delivery_date_filter(picking_pool, delivery_range)
        df = merge_and_clean(picking_pool, sku_master)
        df = classify_and_assign(df)

        # Assign job numbers with the advanced logic
        df = assign_job_numbers_with_scenarios(df)

        output_df = finalize_output(df, gi_type)

        st.success("✅ Processing complete!")
        st.dataframe(output_df.head(20))

        output = export_to_excel(output_df)

        now = datetime.now(ZoneInfo("Asia/Singapore"))
        timestamp = now.strftime("%d %b - %H%M")
        filename = f"master_pick_ticket_{timestamp}.xlsx"

        st.download_button(
            label="Download Master Pick Ticket",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.session_state["final_df"] = output_df

    except Exception as e:
        st.error(f"❌ Error during processing: {e}")

# ------------------------ EXECUTE ------------------------
if picking_pool_file and sku_master_file:
    main()
else:
    st.info("👈 Please upload both Picking Pool and SKU Master Excel files to begin.")
